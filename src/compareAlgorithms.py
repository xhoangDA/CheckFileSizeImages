import os
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import smbclient
import magic
import datetime
import json
import re
import datetime
import sys

def log(message):
    current_time = datetime.datetime.now()
    formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{formatted_time}] {message}"
    print(log_message)

def connectSMB(filePath):
    with open(filePath, 'r') as config_file:
        config_data = json.load(config_file)
    smb_username = config_data["username"]
    smb_password = config_data["password"]
    try:
        smbclient.ClientConfig(username=smb_username, password=smb_password)
    except Exception as e:
        log("\tERROR: Kết nối SMB thất bại. ❌")
        print("==> Error detail: {e}")
 
# function to get all files recursively on the diretory 
def getFiles(dir_path):
    listFiles = []
    try:
        for (path, subdirs, files) in os.walk(dir_path):
            for name in files:
                filePath = os.path.join(path, name)
                relativePathDir = filePath.replace(dir_path,"")
                pathWithoutFilename = relativePathDir.replace(name,"")
                # hàm lấy giá trị filesize (đơn vị byte)
                fileSize = os.stat(filePath).st_size
                fileExt = fileExtension(filePath)
                fileType = detectHeaderFile(filePath)
                element = [pathWithoutFilename, name, relativePathDir, fileSize, fileExt, fileType]
                listFiles.append(element)
    except FileNotFoundError:
        print(f"The directory {dir_path} does not exist")
    except PermissionError:
        print(f"Permission denied to access the directory {dir_path}")
    except OSError as e:
        print(f"An OS error occurred: {e}")
    return listFiles
# Note: Chưa xử lý đc chỗ điều kiện exception

def compareList(list1, list2):
    result = []
    serial = 0
    cloneList1 = list1.copy()
    cloneList2 = list2.copy()
    for i in cloneList2:
        for j in cloneList1:
            if i[2] == j[2]:
                # So sánh file trước và file sau
                changedSize = i[3]-j[3]
                # Nếu dung lượng không thay đổi -> bỏ qua
                if changedSize == 0:
                    cloneList1.remove(j)
                    break
                # Nếu dung lượng file thay đổi -> lưu bản ghi
                else:
                    # Kiểm tra filetype
                    status = "Sửa"
                    serial += 1
                    plus = ""
                    oldSizeKB = round (j[3] / 1024, 2)
                    newSizeKB = round (i[3] / 1024, 2)
                    checkSize = checkFilesWithLargeFilesize(status, changedSize, oldSizeKB)
                    checkType = checkFileType(i[4],i[5])
                    # Đổi đơn vị từ byte thành KB
                    if checkSize != "" and checkType != '': plus = ' + '
                    note = checkType + plus + checkSize
                    element =  [serial, j[2], status, oldSizeKB, newSizeKB, changedSize, "Thay đổi thông tin", note, i[4], i[5]]
                    # element =  [serial, j[2], status, oldSizeKB, newSizeKB, changedSize, "Thay đổi thông tin", note]
                    result.append(element)
                    cloneList1.remove(j)
                    break
            else:
                if j == cloneList1[-1]:
                    # Kiểm tra filetype
                    checkType = checkFileType(i[4],i[5])
                    serial += 1
                    status = "Thêm mới"
                    plus = ""
                    filesizeKB = round (i[3] / 1024, 1)
                    checkSize = checkFilesWithLargeFilesize(status, i[3], 0)
                    if checkSize != "" and checkType != '': plus = ' + '
                    note = checkType + plus + checkSize
                    # note = checkFilesWithLargeFilesize(status, i[3], None)
                    element =  [serial, i[2], status, None, filesizeKB, i[3], "Tạo mới", note, i[4], i[5]]
                    result.append(element)
                    break
                else:
                    continue
    if cloneList1 != [['', '', '', 0, '', '']]:
        for k in cloneList1:
            serial += 1
            filesizeKB = round (k[3] / 1024, 1)
            element =  [serial, '' + k[2], "Xóa", filesizeKB, None, -k[3], "Không còn sử dụng", '', k[4], k[5]]
            result.append(element)
    return result

def countFiles(listFiles):
    return str(len(listFiles))

def totalSize(listFiles):
    total = 0
    for i in listFiles:
        total += i[3]
    totalMB = round(total / (1024*1024), 4)
    return str(totalMB)

# Function check files with large filesize
def checkFilesWithLargeFilesize(status, sizeChange, oldSize):
    warning = ""
    sizeChangeKB = round(sizeChange / 1024, 2)
    if status == "Sửa":
            # If the file already exists, check if the file size after editing is more than 10KB, then give a warning
            if sizeChangeKB > 10:
                warning = "File sau khi chỉnh sửa có kích thước lớn hơn nhiều so với file cũ: " + '{:.2f}'.format(sizeChangeKB) + " (KB)"
    return warning

def writeToExcelFile(filesDir1, filesDir2, resultList, image1, image2, productName, version):
    curentDate = datetime.date.today().strftime("%d%m%Y")
    wb = openpyxl.Workbook()
    # wb = openpyxl.load_workbook(outputFile)
    ws = wb.active
    ws["A1"] = "BẢNG KHAI BÁO CẤU TRÚC VÀ DUNG LƯỢNG PHÁT HÀNH SẢN PHẨM: " + productName
    ws.merge_cells("A1:H1")
    ws["A2"] = "Tên image:"
    ws["B2"] = f"Old image: {image1}"
    if filesDir1 == [['', '', '', 0, '', '']]:
        ws["F2"] = "Số files: 0 - Dung lượng: " + totalSize(filesDir1) + "MB"
    else:
        ws["F2"] = "Số files: " +  countFiles(filesDir1) + " - Dung lượng: " + totalSize(filesDir1) + "MB"
    ws["B3"] = f"New image: {image2}"
    ws["F3"] = "Số files: " +  countFiles(filesDir2) + " - Dung lượng: " + totalSize(filesDir2) + "MB"
    # fieldNames = ["STT", version, "Tình trạng", "Dung lượng cũ (KB)", "Dung lượng mới (KB)", "Chênh lệch (B)", "Mục đích sử dụng", "Ghi chú"]
    fieldNames = ["STT", version, "Tình trạng", "Dung lượng cũ (KB)", "Dung lượng mới (KB)", "Chênh lệch (B)", "Mục đích sử dụng", "Ghi chú", "File extension", "Header file"]
    ws.append(fieldNames)
    for k in resultList:
        ws.append(k)

    # Create a few styles
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    big_bold_blue_font = Font(color = '173B9E', bold=True, size = 14)
    double_border_side = Side(border_style='thin')
    square_border = Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)
    bold12 = Font(bold=True, size = 12)
    gray_background = PatternFill('solid', start_color = 'B1B1B2')
    pink_background = PatternFill('solid', start_color = 'FFCBCB')
    yellow_background = PatternFill('solid', start_color = 'FFFFD1')

    #  Decorate cell A1-A2
    ws["A1"].font = big_bold_blue_font
    ws["A1"].alignment = center_aligned_text
    ws["A2"].font = bold12
    ws.row_dimensions[4].height = 28

    #  Decorate cell 4
    for cell in ws[4]:
        cell.font = bold12
        cell.fill = gray_background
        cell.border = square_border
        cell.alignment = center_aligned_text

    for i in range (4, ws.max_row+1):
        ws["A"+ str(i)].alignment = center_aligned_text
    
    # Highligh suspicious files
    for i in range (5, ws.max_row+1):
        if ws['H' + str(i)].value == 'Không phân tích được file' or ws['H' + str(i)].value == 'File text':
            for cell in ws[i]:
                cell.fill = yellow_background
        elif ws['H' + str(i)].value != '':
            for cell in ws[i]:
                cell.fill = pink_background

    # Custom column width size
    ws.column_dimensions["A"].width = len(ws["A2"].value) + 2

    columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for index, column_letter in enumerate(columns):
        values = [ws.cell(row=i,column=index+2).value for i in range(1,ws.max_row+1)]
        max_width = 0
        for value in values:
            if value == None: continue
            elif max_width <= len(str(value)): max_width = len(value) + 8
        ws.column_dimensions[str(column_letter)].width = max_width
    
    # Chuẩn hóa lại tên image để đưa vào tên file excel
    normalizeImage = str(image2).replace("192.168.41.109/", "").replace("192.168.101.51/", "")
    normalizeImage = re.sub('[\/\\-:\s]+', '_', normalizeImage)

    # Tạo thư mục lưu trữ output và xuất file excel
    os.system(f"mkdir -p /tmp/checkfilesize/output/{productName}")
    filename = '/tmp/checkfilesize/output/%s/%s_%s_%s.xlsx' % (productName, normalizeImage, version, curentDate)
    # filename = '%s_%s_%s.xlsx' % (normalizeImage, version, curentDate)
    wb.save(filename)

    return filename

def detectHeaderFile(pathToFile):
    fileType = magic.from_file(pathToFile)
    # with open(r""+pathToFile, mode='r',encoding='latin1') as file:
    #     data = file.read()
    #     data = bytes(data, 'latin1')
    #     fileType = magic.from_buffer(data)
    return fileType

def fileExtension(pathToFile):
    split_tup = os.path.splitext(pathToFile)
    return split_tup[1]

def checkPath(filePath, version):
    x = filePath.split('\\')
    versionPath = str(x[-1]).lower().replace('.','')
    version = str(version).lower().replace('.','')
    if version in versionPath:
        return True
    else:
        return False
    
def saveExcelToSMB(src, desPath, productName):
    curent_year = datetime.date.today().strftime("%Y")
    curent_month = datetime.date.today().strftime("%m")
    listDir1 = smbclient.listdir(desPath)
    nomarlizeSrc = src.split('/')[-1]
    
    try:
        with open(src, 'rb') as f1:
            content = f1.read()
        # List tất cả các thư mục bên trong thư mục hiện tại
        if curent_year not in listDir1:
            smbclient.mkdir(r"" + desPath + "\\" + curent_year)
        listDeepPath1 = smbclient.listdir(desPath + "\\" + curent_year)
        if ("Tháng " + curent_month) not in listDeepPath1:
            smbclient.mkdir(r"" + desPath + "\\" + curent_year + "\\" + "Tháng " + curent_month)
        listDeepPath2 = smbclient.listdir(desPath + "\\" + curent_year + "\\" + "Tháng " + curent_month)
        listDeepPath2 = [item.lower() for item in listDeepPath2]
        if productName.lower() not in listDeepPath2:
            smbclient.mkdir(r"" + desPath + "\\" + curent_year + "\\" + "Tháng " + curent_month + "\\" + productName)
        with smbclient.open_file(r"" + desPath + "\\" + curent_year + "\\" + "Tháng " + curent_month + "\\" + productName + "\\" + nomarlizeSrc, mode = 'wb') as f2:
            f2.write(content)
            # os.remove(src)
    except Exception as e:
        log(f"\tERROR: Đẩy file kết quả vào SMB storage thất bại. ❌")
        print(f"==> Error detail: {e}")
        sys.exit(100)
    
def checkFileType(fileExt, fileType):
    codeFileList = ['.js', '.css', '.scss', '.cshtml', '.html', '.xslt', '.aspx', '.ascx'] #
    textList = ['.txt']
    scriptFile = ['.sh', '.bash', '.bat', '.ps1'] #
    excelFile = ['.xls', '.xlsx', '.ods']   #
    docFile = ['.doc', '.docx']         #
    excutableFiles = ['.dll', '.exe']   #
    fontType = ['.eot', '.woff', '.woff2', '.ttf', '.vfb']  #
    xmlFile = ['.xml', '.resx', '.mrt', '.mrz', '.mdc', '.mdz'] #
    imageFile = ['.png', '.svg', '.ico', '.svg', '.gif', '.jpg']  #
    fileType = str(fileType).lower()
    fileExt = str(fileExt).lower()
    if fileType == 'empty':
        return 'Không phân tích được file'
    if fileExt in codeFileList:
        if 'text' in fileType:
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    if fileExt in textList:
        if 'text' in fileType:
            return 'File text'
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'        
    elif fileExt in scriptFile:
        if 'text' in fileType:
            return 'File Script'
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif fileExt in excutableFiles:
        if 'executable' in fileType:
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif fileExt in excelFile:
        if 'excel' in fileType or 'sheet' in fileType or 'microsoft' in fileType or 'document' in fileType:
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif fileExt in docFile:
        if 'document' in fileType or 'microsoft' in fileType:
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif fileExt in fontType:
        if 'font' in fileType or 'opentype' in fileType:
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif fileExt in imageFile:
        if 'image' in fileType or 'icon' in fileType:
            return ''
        elif fileExt == '.png' and fileType == 'data':
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif fileExt in xmlFile:
        if 'xml' in fileType:
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif '.json' in fileExt or '.map' in fileExt:
        if 'json' in fileType or 'text' in fileType:
            return ''
        else:
            return 'File không đúng định dạng (so sánh với đuôi file)'
    elif fileType == 'empty': 
        return 'Không phân tích được file'
    else:
        return 'File có định dạng lạ'
